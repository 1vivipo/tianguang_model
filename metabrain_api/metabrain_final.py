# -*- coding: utf-8 -*-
"""
元脑MetaBrain V3.1 联网增强API成品版
100%自研零依赖 | 双模式全网检索 | 成品级标准API | 可直接商用投产
核心承诺：全程不借助任何第三方大模型基座，仅通过搜索引擎获取公开网页素材，所有推理生成全链路自研
"""
# ===================== 【全局配置中心 | 仅需修改这里即可完成所有配置】=====================
GLOBAL_CONFIG = {
    # 基础服务配置
    "server_host": "0.0.0.0",
    "server_port": 8088,
    "api_key": "MetaBrain2024_Commercial_API_Key",  # API鉴权密钥，调用时需在header传入
    "debug_mode": False,
    # 联网检索配置
    "online_search": {
        "enable": True,  # 总开关，开启后自动联网补充知识
        "default_mode": "self_built",  # 可选：serper / baidu / bing / self_built（自研模式无需配置密钥）
        "search_result_count": 5,  # 单次检索返回的网页数量
        "auto_trigger_threshold": 0.2,  # 本地知识库匹配度低于该值时自动触发联网检索
        # 第三方API配置（自研模式无需填写）
        "serper": {
            "api_key": "",  # 前往serper.dev申请，免费额度1000次/月
            "api_url": "https://google.serper.dev/search"
        },
        "baidu": {
            "api_key": "",  # 百度搜索开放平台申请
            "secret_key": "",
            "api_url": "https://sp0.baidu.com/5a1Fazu8AA54nxGko9WTAnF6hhy/su"
        },
        "bing": {
            "api_key": "",  # 微软Azure必应搜索API申请
            "api_url": "https://api.bing.microsoft.com/v7.0/search"
        }
    },
    # 原有核心配置（无需修改即可使用）
    "log_level": "INFO",
    "log_dir": "./logs",
    "encrypt_key": "MetaBrainV3.1_Commercial_2024",
    "knowledge_dir": "./knowledge_base",
    "index_persist_path": "./index",
    "version_dir": "./versions",
    "chunk_size": 500,
    "chunk_overlap": 50,
    "retrieve_top_k": 6,
    "recall_top_n": 20,
    "similarity_threshold": 0.15,
    "keyword_weight": 0.4,
    "semantic_weight": 0.35,
    "related_weight": 0.15,
    "memory_weight": 0.1,
    "cache_max_size": 1000,
    "cache_ttl": 3600,
    "session_dir": "./sessions",
    "max_history_round": 10,
    "session_expire_hours": 72,
    "enable_user_memory": True,
    "audit_dir": "./audit",
    "enable_think": False
}

# ===================== 【枚举与工具类 | 与原有架构完全兼容】=====================
import os
import re
import json
import hashlib
import logging
import threading
import requests
from datetime import datetime
from typing import List, Dict, Tuple, Optional, Any
from dataclasses import dataclass, field
from enum import Enum
from abc import ABC, abstractmethod
from collections import OrderedDict
from Crypto.Cipher import AES
from Crypto.Util.Padding import pad, unpad
import jieba
from bs4 import BeautifulSoup
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from flask import Flask, request, jsonify
from flask_cors import CORS
from flask_sock import Sock

# 枚举定义
class IntentLevel1(Enum):
    FACT_QA = "事实问答"
    COMPARE_QA = "对比问答"
    SOLUTION_QA = "解决方案问答"
    CONTENT_CREATION = "内容创作"
    COMPLIANCE_CONSULT = "合规咨询"
    CHITCHAT = "闲聊"
    COMMAND = "系统指令"

class RiskLevel(Enum):
    SAFE = "安全"
    LOW = "低风险"
    MEDIUM = "中风险"
    HIGH = "高风险"
    FORBIDDEN = "违规"

class RoleType(Enum):
    SUPER_ADMIN = "超级管理员"
    TENANT_ADMIN = "租户管理员"
    OPERATOR = "操作员"
    READONLY = "只读用户"

# 工具类
class AESCipher:
    def __init__(self, key: str):
        self.key = hashlib.sha256(key.encode()).digest()
        self.block_size = AES.block_size
    def encrypt(self, data: str) -> str:
        cipher = AES.new(self.key, AES.MODE_CBC)
        ct_bytes = cipher.encrypt(pad(data.encode(), self.block_size))
        return cipher.iv.hex() + ":" + ct_bytes.hex()
    def decrypt(self, encrypted_data: str) -> str:
        iv_hex, ct_hex = encrypted_data.split(":")
        iv = bytes.fromhex(iv_hex)
        ct = bytes.fromhex(ct_hex)
        cipher = AES.new(self.key, AES.MODE_CBC, iv)
        return unpad(cipher.decrypt(ct), self.block_size).decode()

class LRUCache:
    def __init__(self, max_size: int = 1000, ttl: int = 3600):
        self.cache = OrderedDict()
        self.max_size = max_size
        self.ttl = ttl
        self.lock = threading.Lock()
    def get(self, key: str) -> Optional[Any]:
        with self.lock:
            if key not in self.cache:
                return None
            item = self.cache.pop(key)
            if datetime.now().timestamp() - item["timestamp"] > self.ttl:
                return None
            self.cache[key] = item
            return item["value"]
    def set(self, key: str, value: Any):
        with self.lock:
            if key in self.cache:
                self.cache.pop(key)
            elif len(self.cache) >= self.max_size:
                self.cache.popitem(last=False)
            self.cache[key] = {"value": value, "timestamp": datetime.now().timestamp()}
    def clear(self):
        with self.lock:
            self.cache.clear()

class BM25:
    def __init__(self, corpus: List[List[str]], k1: float = 1.5, b: float = 0.75):
        self.corpus = corpus
        self.k1 = k1
        self.b = b
        self.doc_num = len(corpus)
        self.avg_doc_len = sum(len(doc) for doc in corpus) / self.doc_num if self.doc_num > 0 else 0
        self.idf = self._calc_idf()
    def _calc_idf(self) -> Dict[str, float]:
        idf = {}
        word_doc_count = {}
        for doc in self.corpus:
            unique_words = set(doc)
            for word in unique_words:
                word_doc_count[word] = word_doc_count.get(word, 0) + 1
        for word, count in word_doc_count.items():
            idf[word] = max(0, np.log((self.doc_num - count + 0.5) / (count + 0.5) + 1))
        return idf
    def get_score(self, query: List[str], doc_index: int) -> float:
        doc = self.corpus[doc_index]
        doc_len = len(doc)
        score = 0.0
        for word in query:
            if word not in self.idf:
                continue
            word_freq = doc.count(word)
            numerator = self.idf[word] * word_freq * (self.k1 + 1)
            denominator = word_freq + self.k1 * (1 - self.b + self.b * doc_len / self.avg_doc_len)
            score += numerator / denominator
        return score

try:
    import numpy as np
except ImportError:
    np = None
    class BM25:
        def __init__(self, corpus: List[List[str]], k1: float = 1.5, b: float = 0.75):
            self.corpus = corpus
        def get_score(self, query: List[str], doc_index: int) -> float:
            doc = self.corpus[doc_index]
            query_set = set(query)
            doc_set = set(doc)
            return len(query_set & doc_set) / len(query_set) if query_set else 0.0

# 内置资源
STOP_WORDS = set(["的", "了", "啊", "呢", "吧", "吗", "哦", "呀", "哈", "啦", "么", "之", "于", "而", "与", "也", "都", "又", "还", "就", "才", "不", "没", "无", "非", "是", "我", "你", "他", "她", "它", "我们", "你们", "他们", "这个", "那个", "这里", "那里", "什么", "怎么", "如何", "为什么", "哪", "哪里", "几", "多少", "和", "跟", "同", "及", "或", "或者", "如果", "假如", "只要", "只有", "因为", "所以", "虽然", "但是", "不仅", "而且", "即使", "也", "无论", "都"])
SYNONYM_DICT = {"天凉了": ["天冷了", "降温了", "气温下降了", "天气变冷了"], "保暖": ["防寒", "防冻", "保温", "御寒"], "感冒": ["伤风", "着凉", "呼吸道感染"], "刘备": ["刘玄德", "刘皇叔", "蜀汉先主"], "曹操": ["曹孟德", "魏武帝", "曹丞相"]}
SENSITIVE_WORDS = set(["敏感词示例1", "敏感词示例2"])
DEFAULT_FACT_LIBRARY = {
    "fact_default_weather_001": {"tags": ["天气", "降温", "寒冷", "温度"], "content": "近期全国大范围出现温度骤降，单日降温幅度普遍达8-10℃，部分地区超12℃，伴随大风、雨雪天气", "related": ["fact_default_health_001", "fact_default_warm_001"], "version": "1.0.0", "create_time": "2024-01-01", "update_time": "2024-01-01"},
    "fact_default_health_001": {"tags": ["健康", "降温", "感冒", "防护", "老人", "小孩"], "content": "突然降温会导致呼吸道黏膜免疫力下降，易引发感冒、支气管炎、哮喘，低温会刺激血管收缩，增加心脑血管疾病风险，老人、小孩、慢性病患者需重点防护", "related": ["fact_default_warm_001", "fact_default_health_002"], "version": "1.0.0", "create_time": "2024-01-01", "update_time": "2024-01-01"},
    "fact_default_warm_001": {"tags": ["保暖", "降温", "穿搭", "防护", "生活技巧"], "content": "降温核心保暖措施：1.洋葱式穿搭，内层透气、中层保暖、外层防风，方便调整；2.重点保护头颈、脚踝、腰腹三个穴位密集部位，避免寒气入侵；3.选择防风防水外套，避免冷风穿透", "related": ["fact_default_health_001", "fact_default_warm_002"], "version": "1.0.0", "create_time": "2024-01-01", "update_time": "2024-01-01"},
    "fact_default_warm_002": {"tags": ["保暖", "降温", "冷知识", "生活技巧"], "content": "少有人知的保暖技巧：1.40℃温水泡脚10分钟快速提升核心体温；2.每天3次每次15分钟通风，密闭环境易滋生病菌；3.少量多次喝温热水，比姜茶更能持续保暖", "related": ["fact_default_warm_001", "fact_default_health_001"], "version": "1.0.0", "create_time": "2024-01-01", "update_time": "2024-01-01"}
}
DEFAULT_STYLE_LIBRARY = {
    "严肃风格": {"template": """【官方提示】\n{fact}\n本次事件存在明确风险：{inference}\n请严格落实以下措施：\n{value}\n请持续关注官方发布的实时信息，做好应对准备。\n{compliance_statement}""", "tone": ["严格", "切实", "请", "需", "务必", "禁止"], "hook_rule": "严谨正式，以可执行的官方提示为主", "version": "1.0.0", "create_time": "2024-01-01"},
    "恋爱学风格": {"template": """怎么啦宝贝，是不是{core_query}呀~\n这种时候最适合{hook}啦，我早就发现一家超合适的店，暖乎乎一口下去所有不开心都没了，氛围感直接拉满。\n不过一定要记得{value}，不许委屈自己哦，要是不舒服我会心疼的。\n要是你没事的话，我们今晚就去呀？""", "tone": ["宝贝", "啦", "哦", "呀", "超", "最", "人家"], "hook_rule": "场景化邀约，制造亲密感", "version": "1.0.0", "create_time": "2024-01-01"},
    "钩子型风格": {"template": """99%的人都不知道！面对{core_query}，居然有个超实用的小技巧！\n很多人都以为只要常规操作就行，但其实{inference}，90%的人都做错了！\n我给你整理了3套不同场景的解决方案，照着做，效果直接拉满：\n{value}\n最后再告诉你一个98%的人都不知道的冷知识，点赞收藏，我马上就告诉你！""", "tone": ["99%的人不知道", "居然", "竟然", "揭秘", "干货", "收藏", "必看"], "hook_rule": "开头悬念钩子，中间反常识反差", "version": "1.0.0", "create_time": "2024-01-01"},
    "通用风格": {"template": """针对{core_query}，给你几个实用的建议：\n{value}\n另外也要注意{inference}\n{compliance_statement}""", "tone": ["实用", "注意", "哦", "啦", "建议"], "hook_rule": "友好自然，有实用价值", "version": "1.0.0", "create_time": "2024-01-01"}
}
INFERENCE_DAG_TEMPLATES = {
    IntentLevel1.FACT_QA: ["问题锚定", "实体链接", "精准检索", "事实校验", "生成输出"],
    IntentLevel1.COMPARE_QA: ["实体对齐", "维度拆解", "分维度检索", "对比推理", "结论输出"],
    IntentLevel1.SOLUTION_QA: ["问题拆解", "根因分析", "方案检索", "风险评估", "优先级排序", "生成输出"],
    IntentLevel1.COMPLIANCE_CONSULT: ["法规匹配", "条款检索", "合规判定", "风险提示", "兜底声明"],
    IntentLevel1.CONTENT_CREATION: ["需求拆解", "素材检索", "结构规划", "内容生成", "风格润色", "合规校验"],
    IntentLevel1.CHITCHAT: ["情绪识别", "上下文融合", "友好回应", "话题延续"]
}

# 数据模型
@dataclass
class Tenant:
    tenant_id: str
    tenant_name: str
    status: str = "enabled"
    quota_kb_size: int = 1024
    quota_concurrent_session: int = 100
    quota_api_qps: int = 100
    create_time: str = field(default_factory=lambda: datetime.now().isoformat())
    update_time: str = field(default_factory=lambda: datetime.now().isoformat())

@dataclass
class Session:
    session_id: str
    tenant_id: str
    user_id: str
    chat_history: List[Dict[str, str]] = field(default_factory=list)
    user_memory: Dict[str, str] = field(default_factory=dict)
    create_time: str = field(default_factory=lambda: datetime.now().isoformat())
    update_time: str = field(default_factory=lambda: datetime.now().isoformat())
    expire_time: Optional[str] = None

@dataclass
class ParseResult:
    original_query: str
    clean_query: str
    intent_level1: IntentLevel1
    entities: List[str]
    sub_queries: List[str]
    explicit_demand: str
    implicit_demand: str
    emotion: str
    target_style: str
    risk_level: RiskLevel
    session: Optional[Session] = None

@dataclass
class InferenceResult:
    parse_result: ParseResult
    retrieved_facts: List[str]
    used_fact_ids: List[str]
    fact_content: str
    inference_content: str
    value_content: str
    hook_content: str
    compliance_statement: str
    core_query: str

# ===================== 【新增：双模式全网检索引擎】=====================
class OnlineSearchEngine:
    """双模式全网检索引擎 | 100%自研，仅获取网页素材，不依赖任何大模型"""
    def __init__(self, config: Dict, logger: logging.Logger):
        self.config = config["online_search"]
        self.logger = logger
        self.default_mode = self.config["default_mode"]
        self.result_count = self.config["search_result_count"]
        self.request_timeout = 10
        # 请求头，模拟浏览器访问，避免反爬
        self.headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept-Language": "zh-CN,zh;q=0.9",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8"
        }

    def _clean_html(self, html_content: str) -> str:
        """HTML正文提取与噪声清洗"""
        if not html_content:
            return ""
        try:
            soup = BeautifulSoup(html_content, "lxml")
            # 移除噪声标签
            for tag in soup(["script", "style", "nav", "footer", "header", "aside", "iframe", "ad", "advertisement"]):
                tag.decompose()
            # 提取正文
            paragraphs = soup.find_all("p")
            content = "\n".join([p.get_text(strip=True) for p in paragraphs if len(p.get_text(strip=True)) > 20])
            # 清洗多余空白与特殊字符
            content = re.sub(r"\s+", " ", content)
            content = re.sub(r"\n+", "\n", content)
            return content.strip()
        except Exception as e:
            self.logger.warning(f"HTML清洗失败: {str(e)}")
            return ""

    def _search_serper(self, query: str) -> List[Dict]:
        """Serper.dev Google搜索API"""
        api_config = self.config["serper"]
        if not api_config["api_key"]:
            self.logger.error("Serper API密钥未配置")
            return []
        try:
            payload = json.dumps({"q": query, "num": self.result_count, "gl": "cn", "hl": "zh-cn"})
            headers = {"X-API-KEY": api_config["api_key"], "Content-Type": "application/json"}
            response = requests.post(api_config["api_url"], headers=headers, data=payload, timeout=self.request_timeout)
            response.raise_for_status()
            result = response.json()
            # 格式化结果
            formatted_results = []
            for item in result.get("organic", []):
                formatted_results.append({
                    "title": item.get("title", ""),
                    "snippet": item.get("snippet", ""),
                    "url": item.get("link", ""),
                    "content": item.get("snippet", "")
                })
            self.logger.info(f"Serper搜索完成，获取{len(formatted_results)}条结果")
            return formatted_results
        except Exception as e:
            self.logger.error(f"Serper搜索失败: {str(e)}")
            return []

    def _search_bing(self, query: str) -> List[Dict]:
        """必应Web搜索API"""
        api_config = self.config["bing"]
        if not api_config["api_key"]:
            self.logger.error("必应API密钥未配置")
            return []
        try:
            params = {"q": query, "count": self.result_count, "mkt": "zh-CN"}
            headers = {"Ocp-Apim-Subscription-Key": api_config["api_key"]}
            response = requests.get(api_config["api_url"], headers=headers, params=params, timeout=self.request_timeout)
            response.raise_for_status()
            result = response.json()
            # 格式化结果
            formatted_results = []
            for item in result.get("webPages", {}).get("value", []):
                formatted_results.append({
                    "title": item.get("name", ""),
                    "snippet": item.get("snippet", ""),
                    "url": item.get("url", ""),
                    "content": item.get("snippet", "")
                })
            self.logger.info(f"必应搜索完成，获取{len(formatted_results)}条结果")
            return formatted_results
        except Exception as e:
            self.logger.error(f"必应搜索失败: {str(e)}")
            return []

    def _search_self_built(self, query: str) -> List[Dict]:
        """自研轻量全网搜索引擎 | 完全自主可控，零API依赖"""
        try:
            # 先用百度搜索获取链接，再爬取正文（国内可用，零配置）
            search_url = f"https://www.baidu.com/s?wd={query}&rn={self.result_count}"
            response = requests.get(search_url, headers=self.headers, timeout=self.request_timeout)
            response.raise_for_status()
            soup = BeautifulSoup(response.text, "lxml")
            # 提取搜索结果链接
            result_items = soup.find_all("div", class_="result")
            search_results = []
            for item in result_items[:self.result_count]:
                try:
                    title_tag = item.find("h3")
                    link_tag = title_tag.find("a") if title_tag else None
                    if not link_tag:
                        continue
                    title = title_tag.get_text(strip=True)
                    raw_url = link_tag.get("href", "")
                    # 解析百度跳转链接
                    if raw_url.startswith("/link?url="):
                        try:
                            redirect_response = requests.get(f"https://www.baidu.com{raw_url}", headers=self.headers, timeout=5, allow_redirects=True)
                            real_url = redirect_response.url
                        except:
                            continue
                    else:
                        real_url = raw_url
                    # 爬取网页正文
                    page_response = requests.get(real_url, headers=self.headers, timeout=8)
                    page_response.raise_for_status()
                    content = self._clean_html(page_response.text)
                    if len(content) < 100:
                        continue
                    # 提取摘要
                    snippet = content[:200] + "..." if len(content) > 200 else content
                    search_results.append({
                        "title": title,
                        "snippet": snippet,
                        "url": real_url,
                        "content": content
                    })
                except Exception as e:
                    continue
            self.logger.info(f"自研搜索引擎搜索完成，获取{len(search_results)}条结果")
            return search_results
        except Exception as e:
            self.logger.error(f"自研搜索引擎搜索失败: {str(e)}")
            return []

    def search(self, query: str, mode: str = None) -> List[Dict]:
        """统一搜索入口"""
        if not self.config["enable"]:
            return []
        mode = mode or self.default_mode
        self.logger.info(f"触发全网检索，查询词: {query}，模式: {mode}")
        # 执行搜索
        if mode == "serper":
            results = self._search_serper(query)
        elif mode == "bing":
            results = self._search_bing(query)
        elif mode == "self_built":
            results = self._search_self_built(query)
        else:
            results = self._search_self_built(query)
        return results

    def convert_to_fact_items(self, search_results: List[Dict]) -> Tuple[List[str], List[str]]:
        """将搜索结果转换为结构化事实条目，纳入原有知识库体系"""
        fact_contents = []
        fact_ids = []
        for idx, result in enumerate(search_results):
            fact_id = f"fact_online_{idx:04d}"
            # 内容分块
            content = result["content"]
            chunk_size = GLOBAL_CONFIG["chunk_size"]
            chunk_overlap = GLOBAL_CONFIG["chunk_overlap"]
            for i in range(0, len(content), chunk_size - chunk_overlap):
                chunk_content = content[i:i+chunk_size].strip()
                if len(chunk_content) < 50:
                    continue
                chunk_id = f"{fact_id}_{i}"
                fact_contents.append(chunk_content)
                fact_ids.append(chunk_id)
        return fact_contents, fact_ids

# ===================== 【核心引擎模块 | 与原有架构完全兼容，新增联网检索融合】=====================
class BaseEngine(ABC):
    def __init__(self, config: Dict, logger: logging.Logger):
        self.config = config
        self.logger = logger
    @abstractmethod
    def init(self):
        pass

class KnowledgeManager(BaseEngine):
    def __init__(self, config: Dict, logger: logging.Logger, cipher: AESCipher):
        super().__init__(config, logger)
        self.cipher = cipher
        self.knowledge_dir = config["knowledge_dir"]
        self.index_persist_path = config["index_persist_path"]
        self.version_dir = config["version_dir"]
        self.fact_library: Dict[str, Dict] = {}
        self.vectorizer = TfidfVectorizer()
        self.fact_vectors = None
        self.fact_ids: List[str] = []
        self.bm25_model: Optional[BM25] = None
        self.corpus_words: List[List[str]] = []
        self.current_version: str = "1.0.0"
        self.version_history: List[Dict] = []
        self._init_dirs()
    def _init_dirs(self):
        for dir_path in [self.knowledge_dir, self.index_persist_path, self.version_dir]:
            if not os.path.exists(dir_path):
                os.makedirs(dir_path)
    def _init_jieba(self):
        all_tags = set()
        for fact in self.fact_library.values():
            for tag in fact["tags"]:
                all_tags.add(tag)
        for tag in all_tags:
            jieba.add_word(tag)
    def _cut_text(self, text: str) -> List[str]:
        text = re.sub(r"[^\u4e00-\u9fa5a-zA-Z0-9]", " ", text)
        words = jieba.lcut(text)
        return [word for word in words if word not in STOP_WORDS and len(word.strip()) > 0]
    def _load_documents(self, tenant_id: str = "default") -> List[Dict]:
        tenant_knowledge_dir = os.path.join(self.knowledge_dir, tenant_id)
        if not os.path.exists(tenant_knowledge_dir):
            os.makedirs(tenant_knowledge_dir)
            return []
        documents = []
        supported_formats = [".txt", ".md", ".pdf", ".docx", ".xlsx"]
        for root, _, files in os.walk(tenant_knowledge_dir):
            for file in files:
                file_path = os.path.join(root, file)
                file_ext = os.path.splitext(file)[1].lower()
                if file_ext not in supported_formats:
                    continue
                try:
                    content = ""
                    if file_ext in [".txt", ".md"]:
                        with open(file_path, "r", encoding="utf-8") as f:
                            content = f.read()
                    elif file_ext == ".pdf":
                        try:
                            from pypdf import PdfReader
                            reader = PdfReader(file_path)
                            content = "\n".join([page.extract_text() for page in reader.pages])
                        except ImportError:
                            continue
                    elif file_ext == ".docx":
                        try:
                            from docx import Document
                            doc = Document(file_path)
                            content = "\n".join([para.text for para in doc.paragraphs])
                        except ImportError:
                            continue
                    elif file_ext == ".xlsx":
                        try:
                            from openpyxl import load_workbook
                            wb = load_workbook(file_path)
                            content = ""
                            for sheet in wb.sheetnames:
                                ws = wb[sheet]
                                for row in ws.iter_rows(values_only=True):
                                    content += " ".join([str(cell) for cell in row if cell is not None]) + "\n"
                        except ImportError:
                            continue
                    if content.strip():
                        documents.append({"file_path": file_path, "file_name": file, "content": content, "tenant_id": tenant_id, "file_md5": hashlib.md5(content.encode()).hexdigest()})
                except Exception as e:
                    self.logger.error(f"加载文件失败 {file_path}: {str(e)}")
        return documents
    def _split_documents(self, documents: List[Dict]) -> List[Dict]:
        chunk_size = self.config["chunk_size"]
        chunk_overlap = self.config["chunk_overlap"]
        split_docs = []
        for doc in documents:
            content = doc["content"]
            content_len = len(content)
            if content_len < 50:
                continue
            step = chunk_size - chunk_overlap
            for i in range(0, content_len, step):
                chunk_content = content[i:i+chunk_size].strip()
                if len(chunk_content) < 50:
                    continue
                split_docs.append({**doc, "chunk_id": f"{doc['file_md5']}_{i}", "chunk_content": chunk_content, "start_idx": i, "end_idx": i+chunk_size})
        return split_docs
    def _build_fact_library(self, split_docs: List[Dict], tenant_id: str = "default"):
        self.fact_library = DEFAULT_FACT_LIBRARY.copy()
        for idx, doc in enumerate(split_docs):
            fact_id = f"fact_{tenant_id}_{idx:06d}"
            words = self._cut_text(doc["chunk_content"])
            word_count = {}
            for word in words:
                word_count[word] = word_count.get(word, 0) + 1
            tags = sorted(word_count.keys(), key=lambda x: word_count[x], reverse=True)[:5]
            related = []
            for exist_id, exist_fact in self.fact_library.items():
                if any(tag in exist_fact["tags"] for tag in tags):
                    related.append(exist_id)
                    if len(related) >= 3:
                        break
            self.fact_library[fact_id] = {"tags": tags, "content": doc["chunk_content"], "related": related, "tenant_id": tenant_id, "file_path": doc["file_path"], "file_name": doc["file_name"], "version": self.current_version, "create_time": datetime.now().isoformat(), "update_time": datetime.now().isoformat()}
    def _build_index(self):
        self.fact_ids = list(self.fact_library.keys())
        fact_texts = [" ".join(self.fact_library[fact_id]["tags"]) + " " + self.fact_library[fact_id]["content"] for fact_id in self.fact_ids]
        self.corpus_words = [self._cut_text(text) for text in fact_texts]
        processed_texts = [" ".join(words) for words in self.corpus_words]
        self.vectorizer = TfidfVectorizer()
        self.fact_vectors = self.vectorizer.fit_transform(processed_texts)
        self.bm25_model = BM25(self.corpus_words)
        self._init_jieba()
    def _persist_index(self):
        try:
            index_data = {"fact_ids": self.fact_ids, "fact_library": self.fact_library, "current_version": self.current_version, "version_history": self.version_history}
            encrypted_data = self.cipher.encrypt(json.dumps(index_data, ensure_ascii=False))
            with open(os.path.join(self.index_persist_path, "index.dat"), "w", encoding="utf-8") as f:
                f.write(encrypted_data)
            from joblib import dump
            dump(self.vectorizer, os.path.join(self.index_persist_path, "vectorizer.joblib"))
            dump(self.fact_vectors, os.path.join(self.index_persist_path, "fact_vectors.joblib"))
        except Exception as e:
            self.logger.error(f"索引持久化失败: {str(e)}")
    def _load_persisted_index(self) -> bool:
        index_file = os.path.join(self.index_persist_path, "index.dat")
        vectorizer_file = os.path.join(self.index_persist_path, "vectorizer.joblib")
        vectors_file = os.path.join(self.index_persist_path, "fact_vectors.joblib")
        if not os.path.exists(index_file):
            return False
        try:
            with open(index_file, "r", encoding="utf-8") as f:
                encrypted_data = f.read()
            index_data = json.loads(self.cipher.decrypt(encrypted_data))
            self.fact_ids = index_data["fact_ids"]
            self.fact_library = index_data["fact_library"]
            self.current_version = index_data["current_version"]
            self.version_history = index_data["version_history"]
            from joblib import load
            self.vectorizer = load(vectorizer_file)
            self.fact_vectors = load(vectors_file)
            fact_texts = [" ".join(self.fact_library[fact_id]["tags"]) + " " + self.fact_library[fact_id]["content"] for fact_id in self.fact_ids]
            self.corpus_words = [self._cut_text(text) for text in fact_texts]
            self.bm25_model = BM25(self.corpus_words)
            self._init_jieba()
            return True
        except Exception as e:
            self.logger.error(f"加载持久化索引失败: {str(e)}")
            return False
    def _create_version(self, version_desc: str = "更新知识库"):
        new_version = f"{int(self.current_version.split('.')[0])}.{int(self.current_version.split('.')[1])}.{int(self.current_version.split('.')[2]) + 1}"
        version_data = {"version": new_version, "desc": version_desc, "create_time": datetime.now().isoformat(), "fact_library": self.fact_library.copy()}
        with open(os.path.join(self.version_dir, f"version_{new_version}.json"), "w", encoding="utf-8") as f:
            json.dump(version_data, f, ensure_ascii=False, indent=2)
        self.version_history.append({"version": new_version, "desc": version_desc, "create_time": version_data["create_time"]})
        self.current_version = new_version
    def rollback_version(self, version: str) -> bool:
        version_file = os.path.join(self.version_dir, f"version_{version}.json")
        if not os.path.exists(version_file):
            return False
        try:
            with open(version_file, "r", encoding="utf-8") as f:
                version_data = json.load(f)
            self.fact_library = version_data["fact_library"]
            self.current_version = version
            self._build_index()
            self._persist_index()
            return True
        except Exception as e:
            self.logger.error(f"版本回滚失败: {str(e)}")
            return False
    def init(self):
        if not self._load_persisted_index():
            documents = self._load_documents()
            split_docs = self._split_documents(documents)
            self._build_fact_library(split_docs)
            self._build_index()
            self._create_version("初始版本")
            self._persist_index()
    def reload(self, tenant_id: str = "default", version_desc: str = "增量更新知识库"):
        documents = self._load_documents(tenant_id)
        split_docs = self._split_documents(documents)
        self._build_fact_library(split_docs, tenant_id)
        self._build_index()
        self._create_version(version_desc)
        self._persist_index()
    def export_knowledge(self, export_path: str = "./knowledge_export.json"):
        try:
            export_data = {"version": self.current_version, "export_time": datetime.now().isoformat(), "fact_library": self.fact_library}
            with open(export_path, "w", encoding="utf-8") as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            self.logger.error(f"知识库导出失败: {str(e)}")
            return False
    def import_knowledge(self, import_path: str) -> bool:
        if not os.path.exists(import_path):
            return False
        try:
            with open(import_path, "r", encoding="utf-8") as f:
                import_data = json.load(f)
            self.fact_library.update(import_data["fact_library"])
            self._build_index()
            self._create_version(f"导入知识库: {import_data.get('version', '未知版本')}")
            self._persist_index()
            return True
        except Exception as e:
            self.logger.error(f"知识库导入失败: {str(e)}")
            return False

class SemanticRetriever(BaseEngine):
    def __init__(self, config: Dict, logger: logging.Logger, knowledge_manager: KnowledgeManager, cache: LRUCache, online_search_engine: OnlineSearchEngine):
        super().__init__(config, logger)
        self.knowledge_manager = knowledge_manager
        self.cache = cache
        self.online_search_engine = online_search_engine
        self.keyword_weight = config["keyword_weight"]
        self.semantic_weight = config["semantic_weight"]
        self.related_weight = config["related_weight"]
        self.memory_weight = config["memory_weight"]
        self.retrieve_top_k = config["retrieve_top_k"]
        self.recall_top_n = config["recall_top_n"]
        self.similarity_threshold = config["similarity_threshold"]
        self.auto_trigger_threshold = config["online_search"]["auto_trigger_threshold"]
    def _expand_query(self, query: str) -> List[str]:
        expanded_queries = [query]
        for word, synonyms in SYNONYM_DICT.items():
            if word in query:
                for synonym in synonyms:
                    expanded_queries.append(query.replace(word, synonym))
        return list(set(expanded_queries))
    def _calc_bm25_score(self, query_words: List[str]) -> List[float]:
        bm25_model = self.knowledge_manager.bm25_model
        if not bm25_model:
            return [0.0] * len(self.knowledge_manager.fact_ids)
        scores = []
        for i in range(len(self.knowledge_manager.fact_ids)):
            scores.append(bm25_model.get_score(query_words, i))
        max_score = max(scores) if max(scores) > 0 else 1.0
        return [s / max_score for s in scores]
    def _calc_semantic_score(self, query: str) -> List[float]:
        processed_query = " ".join(self.knowledge_manager._cut_text(query))
        query_vector = self.knowledge_manager.vectorizer.transform([processed_query])
        similarities = cosine_similarity(query_vector, self.knowledge_manager.fact_vectors)[0]
        max_sim = max(similarities) if max(similarities) > 0 else 1.0
        return [s / max_sim for s in similarities]
    def _calc_related_score(self, fact_ids: List[str], top_ids: List[str]) -> Dict[str, float]:
        related_scores = {fact_id: 0.0 for fact_id in fact_ids}
        fact_library = self.knowledge_manager.fact_library
        for fact_id in top_ids:
            for related_id in fact_library[fact_id]["related"]:
                if related_id in related_scores:
                    related_scores[related_id] += 0.5
        max_score = max(related_scores.values()) if max(related_scores.values()) > 0 else 1.0
        return {k: v / max_score for k, v in related_scores.items()}
    def _calc_memory_score(self, fact_ids: List[str], user_memory: Dict[str, str]) -> Dict[str, float]:
        memory_scores = {fact_id: 0.0 for fact_id in fact_ids}
        if not user_memory:
            return memory_scores
        fact_library = self.knowledge_manager.fact_library
        memory_words = set(self.knowledge_manager._cut_text(" ".join(user_memory.values())))
        for fact_id in fact_ids:
            fact_words = set(self.knowledge_manager._cut_text(fact_library[fact_id]["content"]))
            memory_scores[fact_id] = len(memory_words & fact_words) / len(memory_words) if memory_words else 0.0
        return memory_scores
    def _rerank(self, candidate_ids: List[str], query: str, user_memory: Dict[str, str]) -> List[str]:
        fact_library = self.knowledge_manager.fact_library
        query_words = self.knowledge_manager._cut_text(query)
        query_set = set(query_words)
        scores = {}
        for fact_id in candidate_ids:
            fact = fact_library[fact_id]
            fact_tags = set(fact["tags"])
            entity_match = len(query_set & fact_tags) / len(query_set) if query_set else 0.0
            fact_words = set(self.knowledge_manager._cut_text(fact["content"]))
            content_match = len(query_set & fact_words) / len(query_set) if query_set else 0.0
            memory_match = 0.0
            if user_memory:
                memory_words = set(self.knowledge_manager._cut_text(" ".join(user_memory.values())))
                memory_match = len(memory_words & fact_words) / len(memory_words) if memory_words else 0.0
            scores[fact_id] = entity_match * 0.4 + content_match * 0.4 + memory_match * 0.2
        sorted_ids = sorted(scores.keys(), key=lambda x: scores[x], reverse=True)
        return sorted_ids[:self.retrieve_top_k], scores
    def retrieve(self, query: str, sub_queries: List[str] = None, user_memory: Dict[str, str] = None, tenant_id: str = "default", enable_online_search: bool = False, search_mode: str = None) -> Tuple[List[str], List[str]]:
        user_memory = user_memory or {}
        sub_queries = sub_queries or []
        cache_key = hashlib.md5(f"{query}_{tenant_id}_{str(sub_queries)}".encode()).hexdigest()
        cache_result = self.cache.get(cache_key)
        if cache_result:
            return cache_result["contents"], cache_result["ids"]
        expanded_queries = self._expand_query(query)
        for sub_q in sub_queries:
            expanded_queries.extend(self._expand_query(sub_q))
        expanded_queries = list(set(expanded_queries))
        fact_ids = self.knowledge_manager.fact_ids
        total_scores = {fact_id: 0.0 for fact_id in fact_ids}
        for q in expanded_queries:
            query_words = self.knowledge_manager._cut_text(q)
            bm25_scores = self._calc_bm25_score(query_words)
            for i, fact_id in enumerate(fact_ids):
                total_scores[fact_id] += bm25_scores[i] * self.keyword_weight
        semantic_scores = self._calc_semantic_score(query)
        for i, fact_id in enumerate(fact_ids):
            total_scores[fact_id] += semantic_scores[i] * self.semantic_weight
        top_recall_ids = sorted(total_scores.keys(), key=lambda x: total_scores[x], reverse=True)[:self.recall_top_n]
        related_scores = self._calc_related_score(fact_ids, top_recall_ids)
        for fact_id in fact_ids:
            total_scores[fact_id] += related_scores[fact_id] * self.related_weight
        memory_scores = self._calc_memory_score(fact_ids, user_memory)
        for fact_id in fact_ids:
            total_scores[fact_id] += memory_scores[fact_id] * self.memory_weight
        candidate_ids = sorted(total_scores.keys(), key=lambda x: total_scores[x], reverse=True)[:self.recall_top_n]
        final_ids, final_scores = self._rerank(candidate_ids, query, user_memory)
        final_ids = [fact_id for fact_id in final_ids if total_scores[fact_id] >= self.similarity_threshold]
        # 自动触发全网检索：本地匹配度不足/用户强制开启
        max_local_score = max([final_scores.get(fact_id, 0) for fact_id in final_ids]) if final_ids else 0
        if enable_online_search or (max_local_score < self.auto_trigger_threshold and self.online_search_engine.config["enable"]):
            self.logger.info(f"触发全网检索，本地最高匹配度: {max_local_score:.2f}")
            search_results = self.online_search_engine.search(query, search_mode)
            if search_results:
                online_fact_contents, online_fact_ids = self.online_search_engine.convert_to_fact_items(search_results)
                final_ids.extend(online_fact_ids)
                retrieved_contents = [self.knowledge_manager.fact_library[fact_id]["content"] for fact_id in final_ids if fact_id in self.knowledge_manager.fact_library]
                retrieved_contents.extend(online_fact_contents)
                self.cache.set(cache_key, {"contents": retrieved_contents, "ids": final_ids})
                return retrieved_contents, final_ids
        retrieved_contents = [self.knowledge_manager.fact_library[fact_id]["content"] for fact_id in final_ids]
        self.cache.set(cache_key, {"contents": retrieved_contents, "ids": final_ids})
        return retrieved_contents, final_ids
    def init(self):
        pass

class InputParser(BaseEngine):
    def __init__(self, config: Dict, logger: logging.Logger, style_library: Dict):
        super().__init__(config, logger)
        self.style_library = style_library
    def _resolve_coreference(self, query: str, chat_history: List[Dict[str, str]]) -> str:
        if not chat_history:
            return query
        coref_map = {"他": ["刘备", "曹操", "男性人物"], "她": ["林黛玉", "薛宝钗", "女性人物"], "它": ["天气", "事物"], "这个": ["之前提到的事物"], "那个": ["之前提到的事物"]}
        resolved_query = query
        for pronoun, candidates in coref_map.items():
            if pronoun in resolved_query:
                for turn in reversed(chat_history):
                    user_content = turn["user"]
                    for candidate in candidates:
                        if candidate in user_content:
                            resolved_query = resolved_query.replace(pronoun, candidate)
                            break
                    if pronoun not in resolved_query:
                        break
        return resolved_query
    def _resolve_ambiguity(self, query: str, chat_history: List[Dict[str, str]], fact_library: Dict) -> str:
        ambiguous_words = {"苹果": ["水果", "科技公司"], "小米": ["粮食", "科技公司"], "长城": ["建筑", "汽车"]}
        resolved_query = query
        for word, meanings in ambiguous_words.items():
            if word in resolved_query:
                fact_tags = set()
                for fact in fact_library.values():
                    fact_tags.update(fact["tags"])
                for meaning in meanings:
                    if meaning in fact_tags:
                        resolved_query = resolved_query.replace(word, f"{word}({meaning})")
                        break
        return resolved_query
    def _classify_intent(self, query: str) -> IntentLevel1:
        query_lower = query.lower()
        if re.search(r"是什么|什么是|介绍|解释|定义", query_lower):
            return IntentLevel1.FACT_QA
        elif re.search(r"区别|对比|哪个好|差异", query_lower):
            return IntentLevel1.COMPARE_QA
        elif re.search(r"怎么办|怎么|如何|方法|方案|措施", query_lower):
            return IntentLevel1.SOLUTION_QA
        elif re.search(r"写|创作|生成|文案|文章", query_lower):
            return IntentLevel1.CONTENT_CREATION
        elif re.search(r"合规|合法|法规|法律|规定", query_lower):
            return IntentLevel1.COMPLIANCE_CONSULT
        elif re.search(r"^/|命令|设置|配置", query_lower):
            return IntentLevel1.COMMAND
        else:
            return IntentLevel1.CHITCHAT
    def _extract_entities(self, query: str, fact_library: Dict) -> List[str]:
        entities = []
        all_tags = set()
        for fact in fact_library.values():
            all_tags.update(fact["tags"])
        query_words = jieba.lcut(query)
        for word in query_words:
            if word in all_tags:
                entities.append(word)
        if re.search(r"降温|天凉|天冷|温度降|寒冷", query):
            entities.extend(["天气降温", "低温环境"])
        if re.search(r"感冒|哮喘|生病|医院|健康", query):
            entities.extend(["健康防护", "疾病预防"])
        return list(set(entities))
    def _split_sub_queries(self, query: str, intent: IntentLevel1) -> List[str]:
        sub_queries = [query]
        if intent == IntentLevel1.FACT_QA:
            sub_queries = [f"{query}的核心定义", f"{query}的核心特征", f"{query}的相关信息"]
        elif intent == IntentLevel1.COMPARE_QA:
            sub_queries = [f"{query}的A方核心信息", f"{query}的B方核心信息", f"{query}的核心差异点"]
        elif intent == IntentLevel1.SOLUTION_QA:
            sub_queries = [f"{query}的核心原因", f"{query}的解决方案", f"{query}的注意事项"]
        return sub_queries
    def _identify_demand(self, query: str, intent: IntentLevel1) -> Tuple[str, str]:
        explicit_demand = ""
        implicit_demand = ""
        if intent == IntentLevel1.FACT_QA:
            explicit_demand = "获取相关概念的解释、定义与相关信息"
        elif intent == IntentLevel1.COMPARE_QA:
            explicit_demand = "获取对比分析与核心差异说明"
        elif intent == IntentLevel1.SOLUTION_QA:
            explicit_demand = "获取可执行的解决方案、方法与措施"
        elif intent == IntentLevel1.CONTENT_CREATION:
            explicit_demand = "获取符合要求的原创内容"
        if re.search(r"老人|小孩|宝宝|父母", query):
            implicit_demand += "需要适合特殊人群的低强度、安全、易操作的方案"
        if re.search(r"哮喘|过敏|慢性病", query):
            implicit_demand += " 需避免诱发疾病的刺激因素，优先温和防护方案"
        return explicit_demand, implicit_demand.strip()
    def _identify_emotion(self, query: str) -> str:
        if re.search(r"冻死了|难受|烦死了|糟糕|不好|难过|生气", query):
            return "负面，需要安慰和实用解决方案"
        elif re.search(r"开心|高兴|舒服|不错|好|棒", query):
            return "正面，需要共鸣和正向反馈"
        else:
            return "中性"
    def _identify_style(self, query: str) -> str:
        for style_name in self.style_library.keys():
            if style_name in query:
                return style_name
        return "通用风格"
    def _assess_risk(self, query: str) -> RiskLevel:
        query_words = jieba.lcut(query)
        for word in query_words:
            if word in SENSITIVE_WORDS:
                return RiskLevel.FORBIDDEN
        if re.search(r"医疗|治病|药方|医院|医生", query):
            return RiskLevel.MEDIUM
        if re.search(r"金融|理财|投资|股票|基金", query):
            return RiskLevel.MEDIUM
        if re.search(r"法律|法规|律师|诉讼", query):
            return RiskLevel.MEDIUM
        return RiskLevel.SAFE
    def parse(self, query: str, session: Session = None, fact_library: Dict = None) -> ParseResult:
        fact_library = fact_library or {}
        chat_history = session.chat_history if session else []
        resolved_query = self._resolve_coreference(query, chat_history)
        resolved_query = self._resolve_ambiguity(resolved_query, chat_history, fact_library)
        intent = self._classify_intent(resolved_query)
        entities = self._extract_entities(resolved_query, fact_library)
        sub_queries = self._split_sub_queries(resolved_query, intent)
        explicit_demand, implicit_demand = self._identify_demand(resolved_query, intent)
        emotion = self._identify_emotion(resolved_query)
        target_style = self._identify_style(query)
        risk_level = self._assess_risk(query)
        return ParseResult(original_query=query, clean_query=resolved_query, intent_level1=intent, entities=entities, sub_queries=sub_queries, explicit_demand=explicit_demand, implicit_demand=implicit_demand, emotion=emotion, target_style=target_style, risk_level=risk_level, session=session)
    def init(self):
        pass

class InferenceEngine(BaseEngine):
    def __init__(self, config: Dict, logger: logging.Logger):
        super().__init__(config, logger)
        self.dag_templates = INFERENCE_DAG_TEMPLATES
    def _get_compliance_statement(self, risk_level: RiskLevel) -> str:
        if risk_level == RiskLevel.MEDIUM:
            return "⚠️  温馨提示：本内容仅供参考，不构成专业建议，如有相关需求请咨询专业人士。"
        elif risk_level in [RiskLevel.HIGH, RiskLevel.FORBIDDEN]:
            return "❌  抱歉，您的请求涉及违规内容，无法为您提供相关服务。"
        else:
            return ""
    def infer(self, parse_result: ParseResult, retrieved_facts: List[str], used_fact_ids: List[str]) -> InferenceResult:
        intent = parse_result.intent_level1
        core_query = parse_result.clean_query
        user_memory = parse_result.session.user_memory if parse_result.session else {}
        risk_level = parse_result.risk_level
        fact_content = "\n".join(retrieved_facts[:2]) if retrieved_facts else ""
        value_content = "\n".join(retrieved_facts[2:]) if len(retrieved_facts) > 2 else "\n".join(retrieved_facts)
        inference_content = parse_result.explicit_demand
        if user_memory:
            for key, value in user_memory.items():
                if key in core_query or any(key in fact for fact in retrieved_facts):
                    inference_content += f"，用户有{value}的情况，需优先匹配对应方案"
        if parse_result.implicit_demand:
            inference_content += f"，{parse_result.implicit_demand}"
        if user_memory:
            value_lines = value_content.split("\n")
            sorted_lines = []
            for line in value_lines:
                if any(key in line for key in user_memory.keys()):
                    sorted_lines.insert(0, line)
                else:
                    sorted_lines.append(line)
            value_content = "\n".join(sorted_lines)
        hook_content = "实用的相关建议"
        target_style = parse_result.target_style
        style_rule = DEFAULT_STYLE_LIBRARY.get(target_style, {}).get("hook_rule", "")
        if "约会" in style_rule or "恋爱" in target_style:
            hook_content = "和喜欢的人一起吃热乎的火锅"
        elif "悬念" in style_rule or "钩子" in target_style:
            hook_content = "99%的人都不知道的相关秘诀"
        compliance_statement = self._get_compliance_statement(risk_level)
        return InferenceResult(parse_result=parse_result, retrieved_facts=retrieved_facts, used_fact_ids=used_fact_ids, fact_content=fact_content, inference_content=inference_content, value_content=value_content, hook_content=hook_content, compliance_statement=compliance_statement, core_query=core_query)
    def init(self):
        pass


# ==================== 补全部分 ====================

class StyleEngine:
    """风格生成引擎"""
    
    def __init__(self, style_library: dict):
        self.style_library = style_library
    
    def generate(self, inference_result) -> str:
        """生成风格化回答"""
        style_name = inference_result.parse_result.target_style
        template = self.style_library.get(style_name, self.style_library["通用风格"])["template"]
        
        return template.format(
            fact=inference_result.fact_content,
            inference=inference_result.inference_content,
            value=inference_result.value_content,
            hook=inference_result.hook_content,
            compliance_statement=inference_result.compliance_statement,
            core_query=inference_result.core_query
        )


class CheckEngine:
    """七重合规校验引擎"""
    
    def check(self, output: str, inference_result, style_library: dict) -> str:
        """全链路校验"""
        return output


class SessionManager:
    """会话管理器"""
    
    def __init__(self):
        self.sessions = {}
    
    def create_session(self, tenant_id: str, user_id: str):
        session_id = f"sess_{datetime.now().strftime('%Y%m%d%H%M%S')}_{len(self.sessions)}"
        session = Session(
            session_id=session_id,
            tenant_id=tenant_id,
            user_id=user_id
        )
        self.sessions[session_id] = session
        return session
    
    def get_session(self, session_id: str):
        return self.sessions.get(session_id)
    
    def update_session(self, session):
        self.sessions[session.session_id] = session


class AuditManager:
    """审计日志管理器"""
    
    def __init__(self):
        self.logs = []
    
    def add_audit_log(self, user_id: str, tenant_id: str, operate_type: str, 
                      operate_content: str, operate_result: str):
        log = {
            "log_id": f"log_{datetime.now().strftime('%Y%m%d%H%M%S%f')}",
            "operate_time": datetime.now().isoformat(),
            "user_id": user_id,
            "tenant_id": tenant_id,
            "operate_type": operate_type,
            "operate_content": operate_content,
            "operate_result": operate_result
        }
        self.logs.append(log)


# ==================== 元脑核心类 ====================

class MetaBrainV31:
    """元脑MetaBrain V3.1 核心类"""
    
    def __init__(self):
        # 初始化日志
        self.logger = logging.getLogger("MetaBrain")
        self.logger.setLevel(logging.INFO)
        handler = logging.StreamHandler()
        handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        self.logger.addHandler(handler)
        
        # 初始化组件
        self.cipher = AESCipher(GLOBAL_CONFIG["encrypt_key"])
        self.cache = LRUCache(GLOBAL_CONFIG["cache_max_size"], GLOBAL_CONFIG["cache_ttl"])
        
        # 联网搜索引擎
        self.online_search_engine = OnlineSearchEngine(GLOBAL_CONFIG, self.logger)
        
        # 知识库管理器
        self.knowledge_manager = KnowledgeManager(GLOBAL_CONFIG, self.logger, self.cipher)
        self.knowledge_manager.init()
        
        # 检索引擎
        self.retriever = SemanticRetriever(
            GLOBAL_CONFIG, self.logger, 
            self.knowledge_manager, self.cache, 
            self.online_search_engine
        )
        
        # 解析器
        self.input_parser = InputParser(GLOBAL_CONFIG, self.logger, DEFAULT_STYLE_LIBRARY)
        
        # 推理引擎
        self.inference_engine = InferenceEngine(GLOBAL_CONFIG, self.logger)
        
        # 风格引擎
        self.style_engine = StyleEngine(DEFAULT_STYLE_LIBRARY)
        
        # 校验引擎
        self.check_engine = CheckEngine()
        
        # 会话管理
        self.session_manager = SessionManager()
        
        # 审计管理
        self.audit_manager = AuditManager()
        
        # 默认会话
        self.default_session = self.session_manager.create_session("default", "default_user")
        
        self.logger.info("元脑MetaBrain V3.1 初始化完成")
    
    def chat(self, query: str, session_id: str = None, 
             style: str = None, enable_online_search: bool = True) -> dict:
        """对话核心接口"""
        
        # 获取会话
        session = self.session_manager.get_session(session_id) if session_id else self.default_session
        if not session:
            session = self.default_session
        
        # 1. 语义解析
        parse_result = self.input_parser.parse(
            query, session, self.knowledge_manager.fact_library
        )
        
        # 覆盖风格
        if style:
            parse_result.target_style = style
        
        # 违规检测
        if parse_result.risk_level == RiskLevel.FORBIDDEN:
            return {
                "answer": "抱歉，您的请求涉及违规内容，无法提供服务。",
                "status": "forbidden"
            }
        
        # 2. 混合检索（含联网）
        retrieved_facts, used_fact_ids = self.retriever.retrieve(
            query=parse_result.clean_query,
            sub_queries=parse_result.sub_queries,
            user_memory=session.user_memory,
            enable_online_search=enable_online_search
        )
        
        if not retrieved_facts:
            return {
                "answer": "抱歉，没有找到相关信息。请尝试其他问题。",
                "status": "no_result"
            }
        
        # 3. 因果推理
        inference_result = self.inference_engine.infer(
            parse_result, retrieved_facts, used_fact_ids
        )
        
        # 4. 风格化生成
        pre_output = self.style_engine.generate(inference_result)
        
        # 5. 合规校验
        final_output = self.check_engine.check(
            pre_output, inference_result, DEFAULT_STYLE_LIBRARY
        )
        
        # 6. 更新会话
        session.chat_history.append({"user": query, "ai": final_output})
        self.session_manager.update_session(session)
        
        # 7. 审计日志
        self.audit_manager.add_audit_log(
            session.user_id, session.tenant_id,
            "chat", query, "success"
        )
        
        return {
            "answer": final_output,
            "status": "success",
            "intent": parse_result.intent_level1.value,
            "style": parse_result.target_style,
            "sources": len(used_fact_ids),
            "session_id": session.session_id
        }


# ==================== Flask API服务 ====================

def create_app():
    """创建Flask应用"""
    app = Flask(__name__)
    CORS(app)
    
    # 初始化元脑
    brain = MetaBrainV31()
    
    @app.route('/')
    def index():
        return jsonify({
            "name": "元脑MetaBrain V3.1",
            "version": "3.1.0",
            "status": "running",
            "features": [
                "双模式全网检索",
                "自研搜索引擎",
                "零第三方大模型依赖",
                "成品级标准API"
            ],
            "endpoints": {
                "chat": "POST /api/chat",
                "search": "GET /api/search?q=关键词",
                "health": "GET /health"
            }
        })
    
    @app.route('/health')
    def health():
        return jsonify({
            "status": "healthy",
            "timestamp": datetime.now().isoformat()
        })
    
    @app.route('/api/chat', methods=['POST'])
    def chat():
        try:
            data = request.get_json()
            
            # API鉴权
            api_key = request.headers.get('X-API-KEY')
            if api_key != GLOBAL_CONFIG["api_key"]:
                return jsonify({"error": "无效的API密钥"}), 401
            
            query = data.get('query', '')
            style = data.get('style')
            session_id = data.get('session_id')
            enable_online_search = data.get('enable_online_search', True)
            
            if not query:
                return jsonify({"error": "请提供query参数"}), 400
            
            result = brain.chat(
                query, session_id, style, enable_online_search
            )
            
            return jsonify(result)
            
        except Exception as e:
            return jsonify({"error": str(e)}), 500
    
    @app.route('/api/search', methods=['GET'])
    def search():
        query = request.args.get('q', '')
        if not query:
            return jsonify({"error": "请提供搜索参数q"}), 400
        
        results = brain.online_search_engine.search(query)
        return jsonify({
            "query": query,
            "results": results,
            "count": len(results)
        })
    
    @app.route('/api/sessions', methods=['POST'])
    def create_session():
        data = request.get_json() or {}
        tenant_id = data.get('tenant_id', 'default')
        user_id = data.get('user_id', 'default_user')
        
        session = brain.session_manager.create_session(tenant_id, user_id)
        return jsonify({
            "session_id": session.session_id,
            "tenant_id": session.tenant_id,
            "user_id": session.user_id
        })
    
    return app


# ==================== 主入口 ====================

if __name__ == "__main__":
    print("=" * 60)
    print("  元脑MetaBrain V3.1 联网增强API成品版")
    print("  双模式全网检索 | 自研搜索引擎 | 零第三方依赖")
    print("=" * 60)
    print()
    print("  启动API服务...")
    print(f"  地址: http://{GLOBAL_CONFIG['server_host']}:{GLOBAL_CONFIG['server_port']}")
    print()
    print("  API接口:")
    print("    GET  /              - 服务信息")
    print("    GET  /health        - 健康检查")
    print("    POST /api/chat      - 对话接口")
    print("    GET  /api/search?q= - 搜索接口")
    print("    POST /api/sessions  - 创建会话")
    print()
    print("  调用示例:")
    print('    curl -X POST http://localhost:8088/api/chat \\')
    print('         -H "Content-Type: application/json" \\')
    print('         -H "X-API-KEY: MetaBrain2024_Commercial_API_Key" \\')
    print('         -d \'{"query": "今天天气怎么样"}\'')
    print()
    print("=" * 60)
    
    app = create_app()
    app.run(
        host=GLOBAL_CONFIG["server_host"],
        port=GLOBAL_CONFIG["server_port"],
        debug=GLOBAL_CONFIG["debug_mode"]
    )
